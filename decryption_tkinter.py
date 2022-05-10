from tkinter import *
from random import *
import openpyxl
import datetime

#this class will have the function for crypting
class __crypt__:                                                       

    def __init__(self,argcount,arga):
        self.count=argcount
        self.a=arga

#this function will create a dictionary with every character linking to one other character   
    def dict_create(self):
        #print(self.count)
        seed(self.count)                                                    #pattern in randomness
        shuffle(self.a)
        value=self.a.copy()
        shuffle(self.a)
        key=self.a.copy()
        final=dict(zip(key,value))                                      #a dictionary is created
        final['\n']='\n'
        self.count=self.count+1                                         #count changes for further coading
        return final                                                                #returning to main() output variable

    #crypting time
    def crypt(self,user,b):                                                 #using the dictionary it will crypt what user entered
        output=""
        for i in user:
            output=output+b[i]
        return output

class __tk__:

    def __init__(selftk,rootarg,countarg,aarg,sheetarg,narg,wbarg):
        selftk.count=countarg
        selftk.sht=sheetarg
        selftk.n=narg
        selftk.a=aarg
        selftk.root=rootarg
        selftk.wb=wbarg

    def func(selftk,event):
        user=selftk.text.get('1.0',END)
        b=selftk.obj2.dict_create()
        selftk.sht.cell(selftk.n+1,1).value=datetime.datetime.now().strftime("%d-%m-%Y")   #save the date
        selftk.sht.cell(selftk.n+1,2).value=datetime.datetime.now().strftime("%H:%M:%S")  #save the time
        selftk.sht.cell(selftk.n+1,3).value=user    #saves what the user originally entered
        output=selftk.obj2.crypt(user,b)                #calling crpyt()
        for i in output:
            selftk.text1.insert(END,i)
        selftk.sht.cell(selftk.n+1,4).value=output
        selftk.wb.save("decrypt.xlsx")                           #saves the file everytime it writes on it
        selftk.n=selftk.n+1

    def reset(selftk,event):
        selftk.text.delete('1.0',END)
        selftk.text1.delete('1.0',END)


    def send(selftk):
        selftk.obj2=__crypt__(selftk.n,selftk.a)
        selftk.sht.cell(selftk.n,1).value="DATE"
        selftk.sht.cell(selftk.n,2).value="TIME"
        selftk.sht.cell(selftk.n,3).value="ORIGINAL"
        selftk.sht.cell(selftk.n,4).value="CRYPTED"
        
        
    def create(selftk):
        selftk.root.geometry('700x700+300+2')
        selftk.root.title('DE-CRYPTION')
        selftk.send()
        f=Frame(selftk.root,bg='yellow',padx=5,pady=5,relief= RIDGE, borderwidth=3)
        f.pack(anchor=N,padx=5,pady=5,fill=BOTH,expand=1)
        f1=Frame(selftk.root,bg='Yellow',padx=5,relief=RIDGE, borderwidth=3)
        f1.pack(anchor=N,padx=5,pady=5,fill=BOTH,expand=1)
        f2=Frame(selftk.root,bg='blue',padx=5,pady=5,relief=RIDGE, borderwidth=3)
        f2.pack(side=TOP,padx=5,pady=5)
        l1=Label(f,text='Please enter your crypted text here: ',padx=5,pady=5,font=('Calibri',10,'bold'),relief=GROOVE, borderwidth=3,bg='#F8A466')
        l1.pack(anchor=NW,padx=15,pady=5)
        l2=Label(f1,text='Your de-crypted text is: ',padx=5,pady=5,font=('Calibri',10,'bold'),relief=GROOVE, borderwidth=3,bg='#F8A466')
        l2.pack(anchor=NW,padx=15,pady=5)
        selftk.text=Text(f,height=11,width=80,bg='#93DFD9',font=('Times new roman',12,'bold'),fg='#0E165F')
        selftk.text.place(relx=0.5,rely=0.5,anchor=CENTER)
        selftk.text1=Text(f1,height=11,width=80,bg='#FF6767',font=('Times new roman',12,'bold'),fg='#520751')
        selftk.text1.place(relx=0.5,rely=0.5,anchor=CENTER)
        b=Button(f2,text='DE-CRYPT',padx=10,pady=10,bg='sky blue',font=('Calibri',11,'bold'))
        b.pack()
        b.bind('<Button-1>',selftk.func)
        b.bind('<Button-3>',selftk.reset)
        selftk.root.bind('<Shift-Return>',selftk.func)
        selftk.root.bind('<Shift-Delete>',selftk.reset)
        root.mainloop()
       


#main
root=Tk()
count =1
a=list(map(chr,map(int,range(32,127))))
wb = openpyxl.load_workbook("E:\\PYTHON\\cryption codes\\tkinter\\decrypt.xlsx")
sheet = wb.active
if(sheet.max_row==1):
    n=sheet.max_row
else:
    n=sheet.max_row+2           #this was done to give a row space when user re-enters data
tk=__tk__(root,count,a,sheet,n,wb)
tk.create()        
